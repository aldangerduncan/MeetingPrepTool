import sys
import os
import io
import json
import urllib.request
from datetime import date, timedelta

# ---------------------------------------------------------------------------
# Configuration — edit SELECTION_CRITERIA to change what gets picked
# ---------------------------------------------------------------------------

PROSPECTOR_REPORTING_DIR = os.path.expanduser("~/prospector_reporting")

SELECTION_CRITERIA = """
You are selecting the most relevant Prospector insights for a B2B sales rep at an advertising intelligence company (IRD Group / Prospector).

Prioritise insights about:
- Account reviews: creative, media, PR or digital agency pitches or reviews
- New campaigns or significant marketing activity from Australian brands
- Contract wins or losses (agency winning or losing a client)
- Companies in advertising, media, or marketing launching new products or expanding
- Notable hires or departures in senior marketing roles

Deprioritise:
- Financial results without a clear marketing angle
- Generic HR/recruitment news unrelated to marketing
- International companies with no Australian presence or relevance
- Minor operational updates (e.g. office moves, minor sponsorships)
"""

# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def get_insight_date_range():
    """Return (from_date, to_date) as YYYY-MM-DD strings.
    On Monday, return last Friday. Otherwise return yesterday."""
    today = date.today()
    if today.weekday() == 0:  # Monday
        target = today - timedelta(days=3)
    else:
        target = today - timedelta(days=1)
    date_str = target.strftime("%Y-%m-%d")
    return date_str, date_str

# ---------------------------------------------------------------------------
# Prospector export
# ---------------------------------------------------------------------------

def fetch_insights_xlsx(from_date, to_date):
    sys.path.insert(0, PROSPECTOR_REPORTING_DIR)
    from prospector_client import ProspectorClient, ProspectorAuthenticationError, ProspectorAPIError
    import local_secrets as secrets

    client = ProspectorClient(
        username=secrets.PROSPECTOR_EMAIL,
        password=secrets.PROSPECTOR_PASSWORD
    )
    return client.export_insights(from_date=from_date, to_date=to_date)

# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

def parse_insights(xlsx_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    def col(row, name):
        try:
            idx = headers.index(name)
            return row[idx].value or ""
        except (ValueError, IndexError):
            return ""

    insights = []
    for row in ws.iter_rows(min_row=2):
        summary = col(row, "InsightSummary")
        if not summary:
            continue
        if "exclusive" in summary.lower() or "🔥" in summary:
            continue
        insights.append({
            "company":        col(row, "Company"),
            "summary":        summary,
            "insight":        col(row, "Insight"),
            "activity_type":  col(row, "Activity type"),
            "industry":       col(row, "Industry"),
            "protags":        col(row, "Protags"),
            "source":         col(row, "Source"),
            "url":            col(row, "URL"),
            "link":           col(row, "Link"),
        })
    return insights

# ---------------------------------------------------------------------------
# AI selection
# ---------------------------------------------------------------------------

def get_openai_key():
    for path in ["../.openai_key", ".openai_key"]:
        if os.path.exists(path):
            with open(path) as f:
                return f.read().strip()
    return None

def select_insights_with_ai(insights, api_key):
    if not insights:
        return []

    # Build a compact list for the prompt
    lines = []
    for i, ins in enumerate(insights):
        lines.append(
            f"[{i}] Company: {ins['company']} | Type: {ins['activity_type']} | "
            f"Industry: {ins['industry']} | Summary: {ins['summary']}"
        )
    insight_list = "\n".join(lines)

    system_prompt = f"""{SELECTION_CRITERIA.strip()}

You will be given a numbered list of insights. Return a JSON array of the indices (numbers) of the 3 best insights to highlight, ordered best-first.

Example response: [4, 17, 2]

Return only the JSON array, nothing else."""

    user_prompt = f"Here are today's Prospector insights:\n\n{insight_list}"

    data = {
        "model": "gpt-4o",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]
    }

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }

    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(data).encode("utf-8"),
        headers=headers
    )

    with urllib.request.urlopen(req) as resp:
        res_json = json.loads(resp.read().decode("utf-8"))
        content = res_json["choices"][0]["message"]["content"].strip()

    indices = json.loads(content)
    return [insights[i] for i in indices if 0 <= i < len(insights)]

# ---------------------------------------------------------------------------
# HTML formatting
# ---------------------------------------------------------------------------

def format_as_html(selected):
    if not selected:
        return '<div class="insight">No Prospector insights found for this period.</div>'

    parts = []
    for ins in selected:
        tag = ins["activity_type"].split()[0] if ins["activity_type"] else ins["industry"] or "Insight"
        text = ins["summary"]
        link = ins["link"] or ins["url"]
        source = ins["source"]

        source_html = ""
        if source and link:
            source_html = f' <a href="{link}" style="color:var(--muted);font-size:0.8rem;" target="_blank">{source}</a>'
        elif link:
            source_html = f' <a href="{link}" style="color:var(--muted);font-size:0.8rem;" target="_blank">View</a>'

        parts.append(
            f'  <div class="insight">\n'
            f'    <span class="tag">{tag}</span>\n'
            f'    {text}{source_html}\n'
            f'  </div>'
        )
    return "\n".join(parts)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    api_key = get_openai_key()
    if not api_key:
        print('<div class="insight">Error: OpenAI key not found.</div>')
        sys.exit(1)

    from_date, to_date = get_insight_date_range()

    try:
        xlsx_bytes = fetch_insights_xlsx(from_date, to_date)
    except Exception as e:
        print(f'<div class="insight">Error fetching Prospector insights: {e}</div>', file=sys.stderr)
        sys.exit(1)

    insights = parse_insights(xlsx_bytes)

    if not insights:
        print('<div class="insight">No insights published for this period.</div>')
        return

    try:
        selected = select_insights_with_ai(insights, api_key)
    except Exception as e:
        # Fall back to first 3 if AI call fails
        print(f"Warning: AI selection failed ({e}), using first 3 insights.", file=sys.stderr)
        selected = insights[:3]

    print(format_as_html(selected))

if __name__ == "__main__":
    main()
